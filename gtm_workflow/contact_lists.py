"""Excel (.xlsx) export + named contact lists for the Find Leads page.

Lists are stored in the DATABASE (ContactList / ContactListItem) so they persist on
hosts with an ephemeral filesystem (e.g. Render free tier). The .xlsx is generated
on the fly for download/export — nothing is written to disk.
"""
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import db, ContactList, ContactListItem

# (Excel header, key on a contact dict). Order defines the sheet column order.
COLUMNS = [
    ("Company", "company"), ("Name", "name"), ("Title", "title"),
    ("Mobile", "mobile"), ("Email", "email"), ("LinkedIn", "linkedin_url"),
    ("City", "city"), ("State", "state"), ("Industry", "industry"),
    ("Company Size", "company_size"), ("Country", "country"),
    ("Source", "source"), ("Apollo ID", "apollo_id"), ("Saved At", "saved_at"),
]
_HEADERS = [h for h, _ in COLUMNS]
# Item columns we copy straight from an incoming contact dict.
_ITEM_FIELDS = ("company", "name", "title", "email", "linkedin_url", "city",
                "state", "industry", "country", "apollo_id")

_HEADER_FILL = PatternFill("solid", fgColor="1A1F36")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def safe_name(name: str) -> str:
    """Human-readable but safe list name."""
    name = (name or "").strip() or "Untitled list"
    name = re.sub(r"[^A-Za-z0-9 _\-]", "", name).strip()
    return (name or "Untitled list")[:80]


def dedup_key(c: dict) -> str:
    """Contact identity for de-dup: email > apollo_id > linkedin > name+company."""
    em = (c.get("email") or "").strip().lower()
    if em:
        return "em:" + em
    aid = (c.get("apollo_id") or "").strip()
    if aid:
        return "ap:" + aid
    ln = (c.get("linkedin_url") or "").strip().lower()
    if ln:
        return "ln:" + ln
    return "nc:" + (c.get("name", "").strip().lower() + "|" + c.get("company", "").strip().lower())


# ─── XLSX generation (works on plain dicts) ───────────────────────────────────

def _mobile_of(c: dict) -> str:
    return (c.get("mobile") or c.get("phone") or "")


def _row_value(c: dict, key: str):
    if key == "mobile":
        return _mobile_of(c)
    if key == "company_size":
        return str(c.get("company_size") or "")
    if key == "source":
        return c.get("source") or "apollo"
    if key == "saved_at":
        v = c.get("saved_at")
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d %H:%M")
        return v or datetime.now().strftime("%Y-%m-%d %H:%M")
    return c.get(key, "")


def _autosize(ws):
    for i, _ in enumerate(COLUMNS, start=1):
        col = ws.cell(row=1, column=i).column_letter
        longest = max([len(str(ws.cell(row=r, column=i).value or "")) for r in range(1, ws.max_row + 1)] or [10])
        ws.column_dimensions[col].width = min(max(longest + 2, 12), 50)


def build_workbook(contacts: list[dict], sheet_title: str = "Contacts") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Contacts")[:31]
    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for c in contacts:
        ws.append([_row_value(c, key) for _, key in COLUMNS])
    _autosize(ws)
    return wb


def to_xlsx_bytes(contacts: list[dict], sheet_title: str = "Contacts") -> io.BytesIO:
    """In-memory .xlsx from contact dicts (used for the immediate 'Download Excel')."""
    buf = io.BytesIO()
    build_workbook(contacts, sheet_title).save(buf)
    buf.seek(0)
    return buf


# ─── DB-backed named lists ────────────────────────────────────────────────────

def _item_to_dict(it: ContactListItem) -> dict:
    return {
        "company": it.company, "name": it.name, "title": it.title,
        "mobile": it.mobile, "email": it.email, "linkedin_url": it.linkedin_url,
        "city": it.city, "state": it.state, "industry": it.industry,
        "company_size": it.company_size, "country": it.country,
        "source": it.source, "apollo_id": it.apollo_id, "saved_at": it.saved_at,
    }


def save_list(name: str, contacts: list[dict], append: bool = False) -> dict:
    """Create a new list (replacing one of the same name) or append to an existing
    one. De-dups within the list. Returns {name, added, duplicates, total}."""
    name = safe_name(name)
    lst = ContactList.query.filter_by(name=name).first()
    if lst and not append:
        # "New list" with a colliding name → start fresh.
        for it in list(lst.items):
            db.session.delete(it)
        db.session.flush()
    if not lst:
        lst = ContactList(name=name)
        db.session.add(lst)
        db.session.flush()

    seen = {it.dedup_key for it in lst.items}
    added, dups = 0, 0
    for c in contacts:
        if not (c.get("name") or c.get("email")):
            continue
        key = dedup_key(c)
        if key in seen:
            dups += 1
            continue
        seen.add(key)
        fields = {f: (c.get(f) or "") for f in _ITEM_FIELDS}
        fields["mobile"] = _mobile_of(c)
        fields["company_size"] = str(c.get("company_size") or "")
        fields["source"] = c.get("source") or "apollo"
        db.session.add(ContactListItem(list_id=lst.id, dedup_key=key, **fields))
        added += 1

    lst.updated_at = datetime.utcnow()
    db.session.commit()
    total = ContactListItem.query.filter_by(list_id=lst.id).count()
    return {"name": name, "added": added, "duplicates": dups, "total": total}


def available_lists() -> list[dict]:
    """All lists with contact counts and last-updated time, newest first."""
    out = []
    for lst in ContactList.query.all():
        out.append({
            "name": lst.name,
            "count": ContactListItem.query.filter_by(list_id=lst.id).count(),
            "modified": (lst.updated_at or lst.created_at or datetime.utcnow()).strftime("%Y-%m-%d %H:%M"),
        })
    out.sort(key=lambda x: x["modified"], reverse=True)
    return out


def list_xlsx_bytes(name: str):
    """An .xlsx of a saved list, or None if the list doesn't exist."""
    lst = ContactList.query.filter_by(name=safe_name(name)).first()
    if not lst:
        return None
    contacts = [_item_to_dict(it) for it in
                ContactListItem.query.filter_by(list_id=lst.id).order_by(ContactListItem.saved_at).all()]
    return to_xlsx_bytes(contacts, sheet_title=lst.name)


def delete_list(name: str) -> bool:
    lst = ContactList.query.filter_by(name=safe_name(name)).first()
    if not lst:
        return False
    db.session.delete(lst)
    db.session.commit()
    return True
